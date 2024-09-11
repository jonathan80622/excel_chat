import pandas as pd
import sys
import openpyxl
import json
from tqdm import tqdm
import pickle
import numpy as np
from openai import OpenAI
import nump
import streamlit as st

wb = openpyxl.load_workbook(sys.argv[1], data_only=False)

def construct_triplet(ws):
    formulae = {}
    numerics = {}
    strings = {}

    import re
    def split_excel_cell(cell):
        match = re.match(r'([A-Z]+)(\d+)', cell)
        if match:
            column, row = match.groups()
            return (column, int(row))

    for row in ws.iter_rows():
        for cell in row:
            # Check if the cell contains a formula
            if cell.data_type == 'f':
                if type(cell.value) == openpyxl.worksheet.formula.ArrayFormula:
                    formulae[cell.coordinate] = cell.value.text
                elif type(cell.value) == str:
                    formulae[cell.coordinate] = cell.value
            elif cell.data_type == 'n':
                if cell.value is not None:
                    numerics[cell.coordinate] = cell.value
            elif cell.data_type == 's':
                if cell.value is not None:
                    strings[cell.coordinate] = cell.value

    return formulae, numerics, strings

def split_into_batches(lst, batch_size):
    return [lst[i:i + batch_size] for i in range(0, len(lst), batch_size)]

a = ['formula', 'numerics','string']
DF = pd.DataFrame(columns=['batch'])

for j in range(2):
    for sheet in [shtnm for shtnm in wb.sheetnames if not shtnm.startswith('_')]:
        batches = split_into_batches([f"{k} {v}" for k,v in construct_triplet(wb[sheet])[j].items()], 30)
        batches = ['; '.join(_) for _ in batches]

        # Generate column names dynamically as a1, a2, a3, ..., an
        columns = [f'{a[j]}_{sheet}' for i in range(len(batches))]

        # Create a DataFrame with a single row, where each cell contains a batch (list of data)
        df = pd.DataFrame(batches, index=columns, columns=['batch'])
        DF = pd.concat([DF,df])
        

def get_oai_embedding(text):
    client = OpenAI(api_key=st.secrets["OPENAI_API_KEY"])
    embedding_response = client.embeddings.create(
        input=text,
        model="text-embedding-3-small"
    )
    return np.array(embedding_response.data[0].embedding)


DF['embedding'] = ''
for i,x in tqdm(enumerate(DF.batch)):
    DF.embedding.iloc[i] = get_oai_embedding(x)

with open('oai_embed_excel.pkl', 'wb') as f:
    pickle.dump(DF,f)